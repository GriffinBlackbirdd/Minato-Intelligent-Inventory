from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse
from pydantic import BaseModel, Field
from pathlib import Path
from typing import List, Optional
import re
import logging
from datetime import datetime
import json
from openpyxl import load_workbook
from agno.agent import Agent
from agno.models.google import Gemini
from agno.media import File

# Configuration
DATA_FOLDER_PATH = "D:\\minato\\Data"  # Hardcoded path to Data folder
TEMPLATES_FOLDER = "templates"  # Folder for invoice templates
INVOICES_FOLDER = "generated_invoices"  # Folder for generated invoices
GOOGLE_API_KEY = "AIzaSyCWznUz8cnPCkzJ6Bu9ikQGWF6kc-ZUu9k"  # Your API key

app = FastAPI(title="Minato Enterprises - Document Processor", version="1.0.0")

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

# Create necessary directories
Path(TEMPLATES_FOLDER).mkdir(exist_ok=True)
Path(INVOICES_FOLDER).mkdir(exist_ok=True)

# Pydantic models
class AadhaarExtraction(BaseModel):
    aadharNumber: str = Field(..., description="Extract the aadhar number from the given pdf")
    address: str = Field(..., description="Extract the full address from the given pdf")

class SearchRequest(BaseModel):
    customer_name: str

class CustomerSuggestion(BaseModel):
    folder_name: str
    full_path: str
    person_name: str
    aadhaar_number: str
    display_text: str  # "Customer Name - 123456789012"

class ProcessRequest(BaseModel):
    folder_path: str

class ExtractionResult(BaseModel):
    success: bool
    aadhar_number: Optional[str] = None
    address: Optional[str] = None
    error: Optional[str] = None

class InvoiceGenerationRequest(BaseModel):
    customer_name: str
    aadhaar_number: str
    address: str
    items: List[dict] = []  # List of items with name, quantity, price
    additional_data: dict = {}  # Any additional data for placeholders

class InvoiceGenerationResult(BaseModel):
    success: bool
    invoice_path: Optional[str] = None
    invoice_number: Optional[str] = None
    download_url: Optional[str] = None
    error: Optional[str] = None

# Initialize the agent
agent = Agent(
    model=Gemini(id="gemini-1.5-flash", api_key=GOOGLE_API_KEY),
    description="You extract Aadhaar number and address information from the PDF file.",
    response_model=AadhaarExtraction,
)

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Invoice counter for unique invoice numbers
invoice_counter_file = Path("invoice_counter.json")

def get_next_invoice_number() -> str:
    """Generate the next invoice number"""
    try:
        if invoice_counter_file.exists():
            with open(invoice_counter_file, 'r') as f:
                data = json.load(f)
                counter = data.get('counter', 0)
        else:
            counter = 0

        counter += 1

        # Save updated counter
        with open(invoice_counter_file, 'w') as f:
            json.dump({'counter': counter}, f)

        # Format: INV-2025-0001
        return f"INV-{datetime.now().year}-{counter:04d}"

    except Exception as e:
        logger.error(f"Error generating invoice number: {e}")
        # Fallback to timestamp-based number
        return f"INV-{datetime.now().strftime('%Y%m%d%H%M%S')}"

def get_template_file() -> Optional[Path]:
    """Find the invoice template file"""
    template_path = Path(TEMPLATES_FOLDER)

    # Look for Excel files in templates folder
    excel_files = list(template_path.glob("*.xlsx")) + list(template_path.glob("*.xls"))

    if excel_files:
        return excel_files[0]  # Return first Excel file found

    return None

def replace_placeholders_in_cell(cell_value: str, placeholder_data: dict) -> str:
    """Replace placeholders in a cell value"""
    if not isinstance(cell_value, str):
        return cell_value

    result = cell_value
    for placeholder, value in placeholder_data.items():
        if placeholder in result:
            result = result.replace(placeholder, str(value))

    return result

def generate_invoice(customer_data: dict, items: List[dict] = None, additional_data: dict = None) -> InvoiceGenerationResult:
    """
    Generate invoice using template and customer data

    Args:
        customer_data: Dictionary containing customer information
        items: List of items for the invoice
        additional_data: Additional data for custom placeholders

    Returns:
        InvoiceGenerationResult: Result of invoice generation
    """
    try:
        # Find template file
        template_file = get_template_file()
        if not template_file:
            return InvoiceGenerationResult(
                success=False,
                error="Invoice template not found. Please add an Excel template to the 'templates' folder."
            )

        logger.info(f"Using template: {template_file}")

        # Generate invoice number and current date
        invoice_number = get_next_invoice_number()
        current_date = datetime.now().strftime("%d/%m/%Y")
        current_datetime = datetime.now().strftime("%d/%m/%Y %H:%M")

        # Prepare placeholder data
        placeholder_data = {
            "{{INVOICE_NUMBER}}": invoice_number,
            "{{DATE}}": current_date,
            "{{DATETIME}}": current_datetime,
            "{{CUSTOMER_NAME}}": customer_data.get('name', ''),
            "{{AADHAAR_NUMBER}}": customer_data.get('aadhaar_number', ''),
            "{{CUSTOMER_ADDRESS}}": customer_data.get('address', ''),
            "{{COMPANY_NAME}}": "Minato Enterprises",
        }

        # Add items data if provided
        if items:
            total_amount = 0
            for i, item in enumerate(items, 1):
                placeholder_data[f"{{{{ITEM_{i}_NAME}}}}"] = item.get('name', '')
                placeholder_data[f"{{{{ITEM_{i}_QUANTITY}}}}"] = item.get('quantity', '')
                placeholder_data[f"{{{{ITEM_{i}_PRICE}}}}"] = item.get('price', '')
                placeholder_data[f"{{{{ITEM_{i}_TOTAL}}}}"] = item.get('total', '')
                total_amount += float(item.get('total', 0))

            placeholder_data["{{TOTAL_AMOUNT}}"] = f"{total_amount:.2f}"

        # Add additional data if provided
        if additional_data:
            for key, value in additional_data.items():
                placeholder_data[f"{{{{{key}}}}}"] = value

        # Load template workbook
        workbook = load_workbook(template_file)
        worksheet = workbook.active

        # Replace placeholders in all cells
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    new_value = replace_placeholders_in_cell(str(cell.value), placeholder_data)
                    if new_value != str(cell.value):
                        cell.value = new_value

        # Generate output filename
        safe_customer_name = re.sub(r'[^\w\s-]', '', customer_data.get('name', 'Customer')).strip()
        safe_customer_name = re.sub(r'[-\s]+', '_', safe_customer_name)
        output_filename = f"Invoice_{safe_customer_name}_{datetime.now().strftime('%Y%m%d')}_{invoice_number.replace('-', '_')}.xlsx"
        output_path = Path(INVOICES_FOLDER) / output_filename

        # Save the generated invoice
        workbook.save(output_path)
        workbook.close()

        logger.info(f"Invoice generated successfully: {output_path}")

        return InvoiceGenerationResult(
            success=True,
            invoice_path=str(output_path),
            invoice_number=invoice_number,
            download_url=f"/download-invoice/{output_filename}",
            error=None
        )

    except Exception as e:
        logger.error(f"Error generating invoice: {str(e)}")
        return InvoiceGenerationResult(
            success=False,
            error=f"Failed to generate invoice: {str(e)}"
        )

def get_customer_suggestions(query: str) -> List[CustomerSuggestion]:
    """
    Get real-time customer suggestions based on partial name input

    Args:
        query (str): Partial customer name query

    Returns:
        List[CustomerSuggestion]: List of matching customer suggestions
    """
    base_path = Path(DATA_FOLDER_PATH)

    if not base_path.exists():
        logger.error(f"Data folder does not exist: {DATA_FOLDER_PATH}")
        return []

    suggestions = []
    search_query = query.lower().strip()

    # Return empty if query is too short
    if len(search_query) < 2:
        return []

    for item in base_path.iterdir():
        if not item.is_dir():
            continue

        # Parse folder name format: "001 Nandu Singh_687480874343"
        folder_parts = item.name.split('_')
        if len(folder_parts) >= 2:
            name_part = '_'.join(folder_parts[:-1])  # Everything except last part (Aadhaar)
            aadhaar_part = folder_parts[-1]  # Last part should be 12-digit Aadhaar

            # Validate Aadhaar number (should be 12 digits)
            if not (aadhaar_part.isdigit() and len(aadhaar_part) == 12):
                continue

            # Extract name after the number (e.g., "001 Nandu Singh" -> "Nandu Singh")
            name_match = re.match(r'^\d{3}\s+(.+)$', name_part)
            if name_match:
                person_name = name_match.group(1)

                # Check if the search query matches the beginning of the person name (case-insensitive)
                if person_name.lower().startswith(search_query):
                    suggestions.append(CustomerSuggestion(
                        folder_name=item.name,
                        full_path=str(item),
                        person_name=person_name,
                        aadhaar_number=aadhaar_part,
                        display_text=f"{person_name} - {aadhaar_part}"
                    ))

    # Sort by person name and limit results
    suggestions.sort(key=lambda x: x.person_name)
    return suggestions[:10]  # Limit to 10 suggestions

def create_uid_filename(name: str) -> str:
    """
    Create UID filename from name (e.g., "Nandu Singh" -> "UID_NANDU_SINGH.pdf")

    Args:
        name (str): Person's name

    Returns:
        str: UID filename
    """
    uid_name = name.replace(' ', '_').upper()
    return f"UID_{uid_name}.pdf"

def find_uid_file(folder_path: Path, expected_filename: str) -> Optional[Path]:
    """
    Find UID file in the folder

    Args:
        folder_path (Path): Path to the folder
        expected_filename (str): Expected UID filename

    Returns:
        Path: Path to UID file if found, None otherwise
    """
    # First try exact match
    exact_path = folder_path / expected_filename
    if exact_path.exists():
        return exact_path

    # If exact match not found, look for any UID_*.pdf file
    uid_files = list(folder_path.glob("UID_*.pdf"))
    if uid_files:
        return uid_files[0]  # Return first UID file found

    return None

def extract_aadhaar_info(pdf_path: Path) -> ExtractionResult:
    """
    Extract Aadhaar information using the agentic model

    Args:
        pdf_path (Path): Path to the PDF file

    Returns:
        ExtractionResult: Extraction results
    """
    try:
        logger.info(f"Processing PDF: {pdf_path}")

        response = agent.run(
            "Please extract the Aadhaar number and complete address from this document.",
            files=[File(filepath=pdf_path)]
        )

        if response.content:
            return ExtractionResult(
                success=True,
                aadhar_number=response.content.aadharNumber,
                address=response.content.address,
                error=None
            )
        else:
            return ExtractionResult(
                success=False,
                aadhar_number=None,
                address=None,
                error='No content in response'
            )

    except Exception as e:
        logger.error(f"Error processing {pdf_path}: {str(e)}")
        return ExtractionResult(
            success=False,
            aadhar_number=None,
            address=None,
            error=str(e)
        )

# API Routes

@app.get("/", response_class=HTMLResponse)
async def read_root():
    """Serve the main HTML page"""
    return FileResponse("static/index.html")

@app.post("/search", response_model=List[CustomerSuggestion])
async def search_customers(request: SearchRequest):
    """
    Get real-time customer suggestions based on partial name
    """
    if not request.customer_name.strip():
        return []

    if len(request.customer_name.strip()) < 2:
        return []

    suggestions = get_customer_suggestions(request.customer_name)

    logger.info(f"Found {len(suggestions)} suggestions for query: {request.customer_name}")

    return suggestions

@app.post("/process", response_model=ExtractionResult)
async def process_customer(request: ProcessRequest):
    """
    Process the selected customer folder to extract Aadhaar information
    """
    folder_path = Path(request.folder_path)

    if not folder_path.exists() or not folder_path.is_dir():
        raise HTTPException(status_code=404, detail="Customer folder not found")

    logger.info(f"Processing customer folder: {folder_path}")

    # Extract person name from folder name
    folder_parts = folder_path.name.split('_')
    if len(folder_parts) >= 2:
        name_part = '_'.join(folder_parts[:-1])
        name_match = re.match(r'^\d{3}\s+(.+)$', name_part)
        person_name = name_match.group(1) if name_match else name_part
    else:
        person_name = folder_path.name

    # Create expected UID filename
    expected_uid_filename = create_uid_filename(person_name)
    logger.info(f"Looking for UID file: {expected_uid_filename}")

    # Find UID file
    uid_file_path = find_uid_file(folder_path, expected_uid_filename)

    if not uid_file_path:
        error_msg = f"UID file not found in customer folder: {folder_path}"
        logger.error(error_msg)
        raise HTTPException(status_code=404, detail=error_msg)

    logger.info(f"Found UID file: {uid_file_path}")

    # Extract information
    result = extract_aadhaar_info(uid_file_path)

    # Print results to terminal
    if result.success:
        print(f"\n{'='*60}")
        print(f"DOCUMENT EXTRACTION SUCCESSFUL")
        print(f"{'='*60}")
        print(f"Customer Folder: {folder_path.name}")
        print(f"Customer Name: {person_name}")
        print(f"UID File: {uid_file_path.name}")
        print(f"Aadhaar Number: {result.aadhar_number}")
        print(f"Address: {result.address}")
        print(f"{'='*60}\n")
    else:
        print(f"\n{'='*60}")
        print(f"DOCUMENT EXTRACTION FAILED")
        print(f"{'='*60}")
        print(f"Customer Folder: {folder_path.name}")
        print(f"Customer Name: {person_name}")
        print(f"Error: {result.error}")
        print(f"{'='*60}\n")

    return result

@app.post("/generate-invoice", response_model=InvoiceGenerationResult)
async def generate_customer_invoice(request: InvoiceGenerationRequest):
    """
    Generate invoice for customer with extracted information
    """
    try:
        customer_data = {
            'name': request.customer_name,
            'aadhaar_number': request.aadhaar_number,
            'address': request.address
        }

        result = generate_invoice(
            customer_data=customer_data,
            items=request.items,
            additional_data=request.additional_data
        )

        # Print invoice generation result to terminal
        if result.success:
            print(f"\n{'='*60}")
            print(f"INVOICE GENERATED SUCCESSFULLY")
            print(f"{'='*60}")
            print(f"Customer: {request.customer_name}")
            print(f"Invoice Number: {result.invoice_number}")
            print(f"File Path: {result.invoice_path}")
            print(f"Download URL: {result.download_url}")
            print(f"{'='*60}\n")
        else:
            print(f"\n{'='*60}")
            print(f"INVOICE GENERATION FAILED")
            print(f"{'='*60}")
            print(f"Customer: {request.customer_name}")
            print(f"Error: {result.error}")
            print(f"{'='*60}\n")

        return result

    except Exception as e:
        logger.error(f"Error in generate_customer_invoice: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate invoice: {str(e)}")

@app.get("/download-invoice/{filename}")
async def download_invoice(filename: str):
    """
    Download generated invoice file
    """
    file_path = Path(INVOICES_FOLDER) / filename

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Invoice file not found")

    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/template-status")
async def check_template_status():
    """
    Check if invoice template is available
    """
    template_file = get_template_file()

    return {
        "template_available": template_file is not None,
        "template_path": str(template_file) if template_file else None,
        "templates_folder": TEMPLATES_FOLDER
    }

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    template_file = get_template_file()

    return {
        "status": "healthy",
        "data_folder": DATA_FOLDER_PATH,
        "templates_folder": TEMPLATES_FOLDER,
        "invoices_folder": INVOICES_FOLDER,
        "template_available": template_file is not None
    }

if __name__ == "__main__":
    import uvicorn

    # Create static directory if it doesn't exist
    Path("static").mkdir(exist_ok=True)

    print(f"Starting Minato Enterprises Document Processor...")
    print(f"Data folder: {DATA_FOLDER_PATH}")
    print(f"Templates folder: {TEMPLATES_FOLDER}")
    print(f"Invoices folder: {INVOICES_FOLDER}")
    print(f"Server will be available at: http://localhost:8000")

    uvicorn.run(app, host="0.0.0.0", port=8000)